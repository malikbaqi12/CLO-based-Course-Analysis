# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║   ResultEngine – Academic Analytics Backend                                ║
# ║   engine.py                                                                ║
# ║   Parses any course result workbook, extracts all data, runs analytics.    ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

import io, re
import numpy as np
import pandas as pd
import openpyxl
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots


# ── Low-level helpers ─────────────────────────────────────────────────────────

def safe_float(v, default=0.0):
    """Any cell value → float. Returns default for None / Excel errors / NaN."""
    if v is None:
        return default
    if isinstance(v, str) and v.strip().startswith("#"):
        return default
    try:
        f = float(v)
        return default if (f != f) else f        # NaN check
    except (ValueError, TypeError):
        return default


def rts(v):
    """Cell → stripped string."""
    return str(v).strip() if v is not None else ""


def col_of(header_row, *keywords):
    """Return index of first header cell containing any keyword."""
    for ci, cell in enumerate(header_row):
        s = rts(cell).lower()
        if any(kw in s for kw in keywords):
            return ci
    return None


def find_sheet(wb, *keywords):
    """Return sheet name whose title contains any keyword (case-insensitive)."""
    for name in wb.sheetnames:
        nl = name.lower()
        if any(k.lower() in nl for k in keywords):
            return name
    return None


# ══════════════════════════════════════════════════════════════════════════════
# MAIN ENGINE CLASS
# ══════════════════════════════════════════════════════════════════════════════

class ResultEngine:
    """
    Parses a result workbook and exposes:
        .meta            – course metadata dict
        .df              – per-student DataFrame
        .stats           – summary statistics dict
        .clo_stats       – CLO attainment dict
        .ga_stats        – GA attainment dict (mapped from Quantized sheet)
        .assessment_avgs – list of {col, label, avg, max}
        .risk_groups     – {at_risk, borderline, top}
        .is_lab          – bool
        .pass_mark       – int
    """

    def __init__(self, file_bytes: bytes, pass_override: int = 60,
                 clo_threshold: int = 50):
        self.pass_override  = pass_override
        self.clo_threshold  = clo_threshold

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        self._wb = wb

        # Locate sheets
        self._sh_main  = find_sheet(wb, "final combined")
        self._sh_clo   = find_sheet(wb, "quantized")
        self._sh_wt    = find_sheet(wb, "weighted theory", "weighted marks")

        if not self._sh_main:
            raise ValueError("'Final Combined Marks Sheet' not found in workbook.")
        if not self._sh_clo:
            raise ValueError("'Quantized Result' sheet not found in workbook.")

        self._ws_main = wb[self._sh_main]
        self._ws_clo  = wb[self._sh_clo]

        # Detect course type from header columns
        hdr = list(self._ws_main.iter_rows(
            min_row=10, max_row=10, values_only=True))[0]
        hdr_str = " ".join(rts(h).lower() for h in hdr if h)
        self.is_lab = any(k in hdr_str for k in
                          ["lab report", "lab project", "lab mid", "lab end",
                           "viva", "open ended"])

        # Run pipeline
        self.meta            = self._extract_meta()
        self.pass_mark       = self._detect_pass_mark()
        self._weights        = self._extract_weights()
        self.df              = self._extract_students()
        self.clo_stats, \
        self.ga_stats        = self._extract_clo_ga()
        self.df              = self._merge_clo_into_df()
        self.stats           = self._compute_stats()
        self.assessment_avgs = self._compute_assessment_avgs()
        self.risk_groups     = self._compute_risks()

    # ─────────────────────────────────────────────────────────────────────────
    # META
    # ─────────────────────────────────────────────────────────────────────────
    def _extract_meta(self) -> dict:
        ws = self._ws_main
        meta_raw = {}
        # Pattern: col0=label, col2=value  (rows 3–9)
        for row in ws.iter_rows(min_row=3, max_row=9, values_only=True):
            if row[0] and row[2]:
                meta_raw[rts(row[0]).lower()] = rts(row[2])
            elif row[0] and row[7]:          # some sheets use col7 for value
                meta_raw[rts(row[0]).lower()] = rts(row[7])

        def gm(*keys, default="N/A"):
            for k in keys:
                for mk, mv in meta_raw.items():
                    if k.lower() in mk:
                        return mv
            return default

        return {
            "code":     gm("course code"),
            "title":    gm("course title"),
            "faculty":  gm("faculty", "name of faculty", "instructor"),
            "semester": gm("offered in", "semester"),
            "program":  gm("program"),
            "batch":    gm("batch"),
            "dept":     gm("department"),
        }

    # ─────────────────────────────────────────────────────────────────────────
    # PASS MARK
    # ─────────────────────────────────────────────────────────────────────────
    def _detect_pass_mark(self) -> int:
        """Scan workbook for 'Having marks less than N' hint."""
        for row in self._ws_main.iter_rows(min_row=3, max_row=12, values_only=True):
            for cell in row:
                s = rts(cell).lower()
                m = re.search(r"less\s+than\s+(\d+)", s)
                if m:
                    return int(m.group(1)) + 1
        return self.pass_override

    # ─────────────────────────────────────────────────────────────────────────
    # WEIGHTS
    # ─────────────────────────────────────────────────────────────────────────
    def _extract_weights(self) -> dict:
        """
        Returns dict of component → max marks.
        Theory default: assignment=10, quiz=10, presentation=15, mid=25, end=40
        Lab default:    lab_reports=10, lab_project=25, mid=25, end=40
        """
        if self.is_lab:
            defaults = {"lab_reports": 10, "lab_project": 25,
                        "lab_presentation": 0, "viva": 0,
                        "open_ended": 0, "lab_quiz": 0,
                        "other": 0, "mid": 25, "end": 40}
        else:
            defaults = {"assignment": 10, "quiz": 10, "presentation": 15,
                        "mid": 25, "end": 40}

        src = self._ws_main
        if self._sh_wt:
            src = self._wb[self._sh_wt]

        for row in src.iter_rows(min_row=3, max_row=12, values_only=True):
            for ci in range(len(row)-1):
                label = rts(row[ci]).lower()
                val   = safe_float(row[ci+1]) or (safe_float(row[ci+2]) if ci+2 < len(row) else 0)
                if not val:
                    continue
                if "assignment" in label and ("weight" in label or "total" in label):
                    defaults["assignment"] = val
                elif "quiz" in label and ("weight" in label or "total" in label):
                    if "lab" not in label:
                        defaults["quiz"] = val
                elif "presentation" in label and "lab" not in label:
                    defaults["presentation"] = val
                elif "mid term" in label and ("weight" in label or "total" in label):
                    defaults["mid"] = val
                elif "end term" in label and ("weight" in label or "total" in label):
                    defaults["end"] = val
                elif "lab report" in label:
                    defaults["lab_reports"] = val
                elif "lab project" in label:
                    defaults["lab_project"] = val

        return defaults

    # ─────────────────────────────────────────────────────────────────────────
    # STUDENT MARKS (Final Combined Sheet)
    # ─────────────────────────────────────────────────────────────────────────
    def _extract_students(self) -> pd.DataFrame:
        ws = self._ws_main
        W  = self._weights

        # Header row = 10, data = row 12 onward (standard for this template)
        HDR_ROW  = 10
        DATA_ROW = 12

        hdr = list(ws.iter_rows(min_row=HDR_ROW, max_row=HDR_ROW, values_only=True))[0]

        # Build column index map
        CI = {}
        CI["sr"]   = col_of(hdr, "sr", "s.no", "sr.")
        CI["reg"]  = col_of(hdr, "regn", "reg", "roll")
        CI["name"] = col_of(hdr, "name")

        if self.is_lab:
            CI["lab_reports"]     = col_of(hdr, "lab report")
            CI["lab_project"]     = col_of(hdr, "lab project")
            CI["lab_presentation"]= col_of(hdr, "lab presentation")
            CI["viva"]            = col_of(hdr, "viva")
            CI["open_ended"]      = col_of(hdr, "open ended")
            CI["lab_quiz"]        = col_of(hdr, "lab quiz")
            CI["other"]           = col_of(hdr, "other activity")
            CI["mid"]             = col_of(hdr, "lab mid", "mid term")
            CI["end"]             = col_of(hdr, "lab end", "end term")
            CI["total"]           = col_of(hdr, "total weighted", "total marks")
            CI["grade"]           = col_of(hdr, "grade")
        else:
            CI["assignment"]   = col_of(hdr, "assg", "assignment")
            CI["quiz"]         = col_of(hdr, "quiz")
            CI["presentation"] = col_of(hdr, "presentation", "project")
            CI["mid"]          = col_of(hdr, "mid term")
            CI["end"]          = col_of(hdr, "end term")
            CI["total"]        = col_of(hdr, "total marks", "theory weighted")
            CI["grade"]        = col_of(hdr, "grade")

        records = []
        for row in ws.iter_rows(min_row=DATA_ROW, max_row=1000, values_only=True):
            sr_val = row[CI["sr"]] if CI["sr"] is not None and CI["sr"] < len(row) else None
            if not isinstance(sr_val, (int, float)):
                continue
            reg  = rts(row[CI["reg"]])  if CI.get("reg")  is not None else ""
            name = rts(row[CI["name"]]) if CI.get("name") is not None else ""
            if not reg or reg.lower() in ("none","nan",""):
                continue

            def g(key, d=0.0):
                ci = CI.get(key)
                if ci is None or ci >= len(row):
                    return d
                return safe_float(row[ci], d)

            total = g("total")
            grade = rts(row[CI["grade"]]) if CI.get("grade") is not None else ""
            pct   = total   # Total marks in Final Combined are already out of 100

            rec = {
                "Sr":         int(sr_val),
                "Reg_No":     reg,
                "Name":       name,
                "TotalMarks": round(pct, 2),
                "Percentage": round(pct, 2),
                "Grade":      grade,
                "Pass_Fail":  "Pass" if pct >= self.pass_mark else "Fail",
            }

            if self.is_lab:
                rec["LabReports"]     = g("lab_reports")
                rec["LabProject"]     = g("lab_project")
                rec["LabPresentation"]= g("lab_presentation")
                rec["Viva"]           = g("viva")
                rec["OpenEnded"]      = g("open_ended")
                rec["LabQuiz"]        = g("lab_quiz")
                rec["OtherActivity"]  = g("other")
                rec["MidTerm"]        = g("mid")
                rec["EndTerm"]        = g("end")
            else:
                rec["Assignment"]  = g("assignment")
                rec["Quiz"]        = g("quiz")
                rec["Presentation"]= g("presentation")
                rec["MidTerm"]     = g("mid")
                rec["EndTerm"]     = g("end")

            records.append(rec)

        return pd.DataFrame(records)

    # ─────────────────────────────────────────────────────────────────────────
    # CLO & GA (Quantized Result Sheet)
    # ─────────────────────────────────────────────────────────────────────────
    def _extract_clo_ga(self):
        """
        Returns:
            clo_data  – dict of reg_no → {CLO 1: pct, CLO 2: pct, …}
            ga_map    – dict of GA_label → col_index (relative to data cols)
        """
        ws = self._ws_clo
        clo_stats = {}    # CLO name → {avg, pass_cnt, fail_cnt, pass_pct}
        ga_stats  = {}    # GA label → avg attainment across all students

        # Find CLO header row
        clo_hdr_row = None
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
            hits = [rts(v) for v in row if v and rts(v).upper().startswith("CLO ")]
            if len(hits) >= 2:
                clo_hdr_row = ri
                break

        if not clo_hdr_row:
            return {}, {}

        clo_hdr = list(ws.iter_rows(
            min_row=clo_hdr_row, max_row=clo_hdr_row, values_only=True))[0]

        # CLO columns: index >= 3, starts with "CLO "
        clo_cols = []
        for ci, v in enumerate(clo_hdr):
            s = rts(v).strip()
            if s.upper().startswith("CLO ") and ci >= 3:
                if s not in [c[1] for c in clo_cols]:
                    clo_cols.append((ci, s))

        # GA columns: just after CLO cols (row below CLO hdr OR skip 1 row)
        # GA labels typically appear 2 rows after CLO header
        ga_hdr_row = clo_hdr_row + 1       # PLO/GA sub-header
        ga_hdr = list(ws.iter_rows(
            min_row=ga_hdr_row, max_row=ga_hdr_row, values_only=True))[0]

        # Some sheets have GA# in cols past the CLO columns
        ga_cols = []
        max_clo_col = max((c[0] for c in clo_cols), default=3)
        for ci, v in enumerate(ga_hdr):
            if ci <= max_clo_col:
                continue
            s = rts(v).strip()
            if re.match(r'^\d+$', s) or re.match(r'^GA\s*\d+', s, re.I):
                ga_label = f"GA{s}" if s.isdigit() else re.sub(r'\s+', '', s.upper())
                ga_cols.append((ci, ga_label))

        # Data rows start 2 rows after CLO header
        data_start = clo_hdr_row + 2
        REG_CI = 1    # always col B
        NAME_CI = 2   # always col C

        # Student-level CLO values
        student_clo: dict = {}    # reg → {clo_name: value}
        student_ga:  dict = {}    # reg → {ga_label: value}

        for row in ws.iter_rows(min_row=data_start, max_row=2000, values_only=True):
            sr = row[0] if row else None
            if not isinstance(sr, (int, float)):
                continue
            reg = rts(row[REG_CI])
            if not reg or reg.lower() in ("none", "nan", ""):
                continue

            clo_vals = {}
            for ci, clo_name in clo_cols:
                if ci < len(row):
                    clo_vals[clo_name] = safe_float(row[ci])

            ga_vals = {}
            for ci, ga_label in ga_cols:
                if ci < len(row):
                    v = safe_float(row[ci])
                    if v > 0:
                        ga_vals[ga_label] = v

            student_clo[reg] = clo_vals
            if ga_vals:
                student_ga[reg] = ga_vals

        # Aggregate CLO stats
        clo_names = [c[1] for c in clo_cols]
        for clo in clo_names:
            vals = [student_clo[r].get(clo, 0) for r in student_clo]
            avg  = round(np.mean(vals), 2) if vals else 0
            pass_cnt = sum(1 for v in vals if v >= self.clo_threshold)
            fail_cnt = len(vals) - pass_cnt
            clo_stats[clo] = {
                "avg":      avg,
                "pass_cnt": pass_cnt,
                "fail_cnt": fail_cnt,
                "pass_pct": round(pass_cnt/len(vals)*100,1) if vals else 0,
                "per_student": {r: student_clo[r].get(clo, 0) for r in student_clo},
            }

        # Aggregate GA stats
        all_gas = set()
        for gv in student_ga.values():
            all_gas.update(gv.keys())
        for ga in sorted(all_gas):
            vals = [student_ga[r].get(ga, 0) for r in student_ga if ga in student_ga.get(r,{})]
            ga_stats[ga] = round(np.mean(vals), 2) if vals else 0

        self._student_clo = student_clo   # cache for merging
        return clo_stats, ga_stats

    # ─────────────────────────────────────────────────────────────────────────
    # MERGE CLO VALUES INTO MAIN DF
    # ─────────────────────────────────────────────────────────────────────────
    def _merge_clo_into_df(self) -> pd.DataFrame:
        df = self.df.copy()
        clo_names = list(self.clo_stats.keys())

        for clo in clo_names:
            per_student = self.clo_stats[clo].get("per_student", {})
            df[clo]          = df["Reg_No"].map(per_student).fillna(0)
            df[f"{clo}_Pass"]= df[clo].apply(
                lambda v: "Pass" if v >= self.clo_threshold else "Fail")

        return df

    # ─────────────────────────────────────────────────────────────────────────
    # SUMMARY STATISTICS
    # ─────────────────────────────────────────────────────────────────────────
    def _compute_stats(self) -> dict:
        df = self.df
        N  = len(df)
        passed = int((df["Pass_Fail"] == "Pass").sum())
        failed = N - passed
        pct = df["Percentage"]
        return {
            "n":        N,
            "passed":   passed,
            "failed":   failed,
            "pass_pct": round(passed/N*100, 1) if N else 0,
            "fail_pct": round(100 - passed/N*100, 1) if N else 0,
            "avg":      round(pct.mean(), 2),
            "high":     round(pct.max(), 2),
            "low":      round(pct.min(), 2),
            "median":   round(pct.median(), 2),
            "std":      round(pct.std(), 2),
        }

    # ─────────────────────────────────────────────────────────────────────────
    # ASSESSMENT AVERAGES
    # ─────────────────────────────────────────────────────────────────────────
    def _compute_assessment_avgs(self) -> list:
        df = self.df
        W  = self._weights

        if self.is_lab:
            mapping = [
                ("LabReports",      "Lab Reports",      W.get("lab_reports",10)),
                ("LabProject",      "Lab Project",      W.get("lab_project",25)),
                ("LabPresentation", "Lab Presentation", W.get("lab_presentation",0)),
                ("Viva",            "Viva Voce",        W.get("viva",0)),
                ("OpenEnded",       "Open Ended",       W.get("open_ended",0)),
                ("MidTerm",         "Mid Term",         W.get("mid",25)),
                ("EndTerm",         "End Term",         W.get("end",40)),
            ]
        else:
            mapping = [
                ("Assignment",   "Assignment",   W.get("assignment",10)),
                ("Quiz",         "Quiz",         W.get("quiz",10)),
                ("Presentation", "Presentation", W.get("presentation",15)),
                ("MidTerm",      "Mid Term",     W.get("mid",25)),
                ("EndTerm",      "End Term",     W.get("end",40)),
            ]

        result = []
        for col, label, max_v in mapping:
            if col in df.columns and df[col].sum() > 0:
                result.append({
                    "col":   col,
                    "label": f"{label} (/{int(max_v)})",
                    "avg":   round(df[col].mean(), 2),
                    "max":   max_v,
                })
        return result

    # ─────────────────────────────────────────────────────────────────────────
    # RISK GROUPS
    # ─────────────────────────────────────────────────────────────────────────
    def _compute_risks(self) -> dict:
        df = self.df
        pm = self.pass_mark

        def row_to_dict(r):
            return {"name": r["Name"], "reg": r["Reg_No"],
                    "pct": r["Percentage"], "grade": r["Grade"]}

        at_risk   = [row_to_dict(r) for _, r in df[df["Pass_Fail"]=="Fail"].iterrows()]
        borderline = [row_to_dict(r) for _, r in
                      df[(df["Pass_Fail"]=="Pass") &
                         (df["Percentage"] < pm+10)].iterrows()]
        top       = [row_to_dict(r) for _, r in
                     df.sort_values("Percentage", ascending=False).head(10).iterrows()
                     if r["Percentage"] >= pm]

        return {"at_risk": at_risk, "borderline": borderline, "top": top}

    # ─────────────────────────────────────────────────────────────────────────
    # HTML DASHBOARD BUILDER
    # ─────────────────────────────────────────────────────────────────────────
    def build_html_dashboard(self, clo_thresh, gc_fn, C, GRADE_ORDER) -> str:
        """Generate the standalone HTML dashboard with all Plotly charts."""
        df         = self.df
        stats      = self.stats
        clo_st     = self.clo_stats
        ga_st      = self.ga_stats
        meta       = self.meta
        assg       = self.assessment_avgs

        grades_present = [g for g in GRADE_ORDER if g in df["Grade"].values]
        grade_dist = df["Grade"].value_counts().reindex(grades_present).dropna().astype(int)
        clo_names  = list(clo_st.keys())
        clo_avgs   = [clo_st[c]["avg"] for c in clo_names]

        all_figs = []

        # Pie charts
        fig1 = make_subplots(1,2, specs=[[{"type":"pie"},{"type":"pie"}]],
                             subplot_titles=["Pass vs Fail","Grade Distribution"])
        fig1.add_trace(go.Pie(
            labels=["Pass","Fail"], values=[stats["passed"],stats["failed"]],
            marker_colors=[C["success"],C["danger"]], hole=0.42,
            textinfo="label+percent+value",
        ),1,1)
        fig1.add_trace(go.Pie(
            labels=list(grade_dist.index), values=list(grade_dist.values),
            marker_colors=[gc_fn(g) for g in grade_dist.index], hole=0.42,
            textinfo="label+value",
        ),1,2)
        fig1.update_layout(height=420, legend=dict(orientation="h",y=-0.1),
                           title=f"<b>Pass/Fail & Grade Overview — {meta['code']}</b>")
        all_figs.append((fig1, "Pass/Fail & Grade Overview"))

        # Grade bar
        fig2 = go.Figure(go.Bar(
            x=list(grade_dist.index), y=list(grade_dist.values),
            marker_color=[gc_fn(g) for g in grade_dist.index],
            text=list(grade_dist.values), textposition="outside",
        ))
        fig2.update_layout(title="<b>Grade Distribution</b>", height=420,
                           plot_bgcolor="#f9f9f9")
        all_figs.append((fig2, "Grade Distribution"))

        # Histogram
        fig3 = go.Figure(go.Histogram(x=df["Percentage"], nbinsx=16,
                                      marker_color=C["primary"], opacity=0.85))
        fig3.add_vline(x=stats["avg"],      line_dash="dash", line_color=C["warning"],
                       annotation_text=f"Avg {stats['avg']}%")
        fig3.add_vline(x=self.pass_mark,    line_dash="dot",  line_color=C["danger"],
                       annotation_text=f"Pass {self.pass_mark}%")
        fig3.update_layout(title="<b>Marks Distribution</b>", height=420,
                           plot_bgcolor="#f9f9f9")
        all_figs.append((fig3, "Marks Distribution"))

        # Assessment
        if assg:
            labels = [a["label"] for a in assg]
            avgs   = [a["avg"]   for a in assg]
            maxs   = [a["max"]   for a in assg]
            pcts   = [round(a/m*100,1) if m else 0 for a,m in zip(avgs,maxs)]
            pct_clrs = [C["success"] if p>=60 else C["warning"] if p>=40
                        else C["danger"] for p in pcts]
            fig4 = make_subplots(1,2,
                subplot_titles=["Avg Obtained vs Max","Performance % per Component"])
            fig4.add_trace(go.Bar(name="Max",x=labels,y=maxs,
                marker_color="#cfd8dc"),1,1)
            fig4.add_trace(go.Bar(name="Avg",x=labels,y=avgs,
                marker_color=C["primary"],text=avgs,textposition="outside"),1,1)
            fig4.add_trace(go.Bar(name="%",x=labels,y=pcts,
                marker_color=pct_clrs,text=[f"{p}%" for p in pcts],
                textposition="outside",showlegend=False),1,2)
            fig4.update_layout(barmode="group", height=420,
                               title="<b>Assessment Component Performance</b>")
            all_figs.append((fig4, "Assessment Performance"))

        # CLO bar
        if clo_st:
            clo_clrs = [C["success"] if v>=clo_thresh else C["danger"] for v in clo_avgs]
            fig5 = go.Figure(go.Bar(
                x=clo_names, y=clo_avgs, marker_color=clo_clrs,
                text=[f"{v:.1f}%" for v in clo_avgs], textposition="outside",
            ))
            fig5.add_hline(y=clo_thresh, line_dash="dash", line_color=C["danger"],
                           annotation_text=f"Threshold {clo_thresh}%")
            fig5.update_layout(title="<b>CLO Attainment</b>", height=420,
                               yaxis=dict(range=[0,110]), plot_bgcolor="#f9f9f9")
            all_figs.append((fig5, "CLO Attainment"))

            # CLO pass/fail
            fig6 = go.Figure()
            fig6.add_trace(go.Bar(name="Pass", x=clo_names,
                y=[clo_st[c]["pass_cnt"] for c in clo_names],
                marker_color=C["success"],
                text=[clo_st[c]["pass_cnt"] for c in clo_names], textposition="inside"))
            fig6.add_trace(go.Bar(name="Fail", x=clo_names,
                y=[clo_st[c]["fail_cnt"] for c in clo_names],
                marker_color=C["danger"],
                text=[clo_st[c]["fail_cnt"] for c in clo_names], textposition="inside"))
            fig6.update_layout(barmode="group", title="<b>CLO Pass/Fail Count</b>",
                               height=420, plot_bgcolor="#f9f9f9")
            all_figs.append((fig6, "CLO Pass/Fail"))

            # Radar
            if len(clo_names) >= 2:
                cats  = clo_names + [clo_names[0]]
                atts  = clo_avgs  + [clo_avgs[0]]
                thr   = [clo_thresh]*len(cats)
                fig7 = go.Figure()
                fig7.add_trace(go.Scatterpolar(r=atts, theta=cats, fill="toself",
                    name="Avg Attainment", line_color=C["primary"],
                    fillcolor="rgba(26,115,232,0.22)"))
                fig7.add_trace(go.Scatterpolar(r=thr, theta=cats, fill="toself",
                    name=f"Threshold ({clo_thresh}%)", line_color=C["danger"],
                    line_dash="dash", fillcolor="rgba(234,67,53,0.07)"))
                fig7.update_layout(title="<b>CLO Radar</b>", height=460,
                    polar=dict(radialaxis=dict(visible=True,range=[0,100])),
                    legend=dict(orientation="h",y=-0.1))
                all_figs.append((fig7, "CLO Radar"))

            # Heatmap
            heat_cols = [c for c in clo_names if c in df.columns]
            if heat_cols:
                heat  = df[["Name"]+heat_cols].set_index("Name")
                fig8  = go.Figure(go.Heatmap(
                    z=heat.values.tolist(), x=heat_cols, y=heat.index.tolist(),
                    colorscale=[[0,"#b71c1c"],[clo_thresh/100,"#fdd835"],[1,"#1b5e20"]],
                    zmin=0, zmax=100,
                    text=[[f"{v:.1f}%" for v in r] for r in heat.values.tolist()],
                    texttemplate="%{text}",
                    colorbar=dict(title="Attainment %",ticksuffix="%"),
                ))
                fig8.update_layout(title="<b>CLO Heatmap (per Student)</b>",
                    height=max(420,20*len(heat)),
                    yaxis=dict(tickfont=dict(size=9)),
                    margin=dict(l=170,r=40))
                all_figs.append((fig8, "CLO Heatmap"))

        # GA bar
        if ga_st:
            ga_names = list(ga_st.keys())
            ga_avgs  = [ga_st[g] for g in ga_names]
            ga_clrs  = [C["success"] if v>=clo_thresh else C["danger"] for v in ga_avgs]
            fig_ga = go.Figure(go.Bar(
                x=ga_names, y=ga_avgs, marker_color=ga_clrs,
                text=[f"{v:.1f}%" for v in ga_avgs], textposition="outside",
            ))
            fig_ga.add_hline(y=clo_thresh, line_dash="dash", line_color=C["danger"],
                             annotation_text=f"Threshold {clo_thresh}%")
            fig_ga.update_layout(title="<b>GA Attainment</b>", height=420,
                                 yaxis=dict(range=[0,115]), plot_bgcolor="#f9f9f9")
            all_figs.append((fig_ga, "GA Attainment"))

        # Top/bottom
        n_show  = min(10, max(1,stats['n']//2))
        sorted_ = df.sort_values("Percentage",ascending=False).reset_index(drop=True)
        top_df  = sorted_.head(n_show)
        bot_df  = sorted_.tail(n_show).sort_values("Percentage")
        fig9 = make_subplots(1,2, subplot_titles=[f"Top {n_show}",f"Bottom {n_show}"])
        fig9.add_trace(go.Bar(x=top_df["Percentage"],y=top_df["Name"],orientation="h",
            marker_color=[gc_fn(g) for g in top_df["Grade"]],
            text=[f"{g}|{p:.1f}%" for g,p in zip(top_df["Grade"],top_df["Percentage"])],
            textposition="inside"),1,1)
        fig9.add_trace(go.Bar(x=bot_df["Percentage"],y=bot_df["Name"],orientation="h",
            marker_color=C["danger"],
            text=[f"{g}|{p:.1f}%" for g,p in zip(bot_df["Grade"],bot_df["Percentage"])],
            textposition="inside"),1,2)
        fig9.update_layout(title="<b>Top & Bottom Students</b>",
            height=max(380,n_show*35+80), showlegend=False, plot_bgcolor="#f9f9f9")
        fig9.update_xaxes(range=[0,105])
        all_figs.append((fig9, "Top & Bottom Students"))

        # Box plot
        fig10 = go.Figure()
        for g in grades_present:
            sub = df[df["Grade"]==g]["Percentage"]
            if not sub.empty:
                fig10.add_trace(go.Box(y=sub, name=g, marker_color=gc_fn(g),
                    boxpoints="all", jitter=0.4, pointpos=-1.6))
        fig10.update_layout(title="<b>Marks Spread by Grade</b>", height=450,
                            showlegend=False, plot_bgcolor="#f9f9f9")
        all_figs.append((fig10, "Box Plot by Grade"))

        # Mid vs End (Theory only)
        if "MidTerm" in df.columns and "EndTerm" in df.columns \
                and df["MidTerm"].sum()>0 and df["EndTerm"].sum()>0:
            fig11 = px.scatter(df, x="MidTerm", y="EndTerm",
                color="Grade", size="Percentage", size_max=22,
                hover_data={"Name":True,"Reg_No":True,"Percentage":":.2f"},
                title="<b>Mid Term vs End Term Scatter</b>")
            fig11.update_layout(height=460, plot_bgcolor="#f9f9f9")
            all_figs.append((fig11, "Mid vs End Scatter"))

        # Assemble HTML
        charts_html = ""
        for i,(fig,name) in enumerate(all_figs):
            charts_html += (
                f'<div class="cw"><h3>{name}</h3>'
                + fig.to_html(full_html=False,
                              include_plotlyjs="cdn" if i==0 else False)
                + "</div>"
            )

        html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="UTF-8">
  <title>{meta['code']} Dashboard</title>
  <style>
    body{{font-family:'Segoe UI',sans-serif;background:#f4f6f9;margin:0;padding:16px}}
    .banner{{background:linear-gradient(135deg,#1a237e,#1565c0);color:#fff;
             border-radius:12px;padding:18px 26px;margin-bottom:20px}}
    .banner h1{{margin:0 0 5px;font-size:22px}}
    .banner p{{margin:0;font-size:12px;opacity:.85}}
    .cw{{background:#fff;border-radius:10px;padding:14px;
         margin-bottom:18px;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
    .cw h3{{font-size:14px;color:#1a237e;margin:0 0 10px}}
    .kpi-row{{display:flex;flex-wrap:wrap;gap:12px;margin-bottom:18px}}
    .kpi{{flex:1 1 120px;background:#fff;border-radius:10px;padding:14px;
           text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
    .kl{{font-size:9px;font-weight:700;color:#555;text-transform:uppercase;letter-spacing:.8px}}
    .kv{{font-size:28px;font-weight:800;margin:4px 0 2px}}
    .ks{{font-size:11px;color:#777}}
  </style>
</head>
<body>
<div class="banner">
  <h1>🎓 {meta['code']} — {meta['title']}</h1>
  <p>Faculty: {meta['faculty']} &nbsp;|&nbsp; Program: {meta['program']}
     &nbsp;|&nbsp; Batch: {meta['batch']} &nbsp;|&nbsp; Semester: {meta['semester']}
     &nbsp;|&nbsp; Type: {"🔬 Lab" if self.is_lab else "📚 Theory"}</p>
</div>
<div class="kpi-row">
  <div class="kpi"><div class="kl">Students</div>
    <div class="kv" style="color:#1a73e8">{stats['n']}</div><div class="ks">enrolled</div></div>
  <div class="kpi"><div class="kl">Passed</div>
    <div class="kv" style="color:#34a853">{stats['passed']}</div>
    <div class="ks">{stats['pass_pct']}% rate</div></div>
  <div class="kpi"><div class="kl">Failed</div>
    <div class="kv" style="color:#ea4335">{stats['failed']}</div>
    <div class="ks">{stats['fail_pct']}% rate</div></div>
  <div class="kpi"><div class="kl">Average</div>
    <div class="kv" style="color:#f9a825">{stats['avg']}%</div>
    <div class="ks">class mean</div></div>
  <div class="kpi"><div class="kl">Highest</div>
    <div class="kv" style="color:#2e7d32">{stats['high']}%</div>
    <div class="ks">top score</div></div>
  <div class="kpi"><div class="kl">Lowest</div>
    <div class="kv" style="color:#c62828">{stats['low']}%</div>
    <div class="ks">bottom score</div></div>
  <div class="kpi"><div class="kl">Std Dev</div>
    <div class="kv" style="color:#9c27b0">{stats['std']}%</div>
    <div class="ks">score spread</div></div>
</div>
{charts_html}
</body>
</html>"""
        return html
